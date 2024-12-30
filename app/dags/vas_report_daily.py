from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.providers.oracle.hooks.oracle import OracleHook
import openpyxl  
from openpyxl.styles import Border, Side
import logging
import os
from datetime import datetime, date, timedelta

logger = logging.getLogger("airflow.task")
logger.setLevel(logging.INFO)
ngay_hien_tai = date.today()
ngay_tong_hop = ngay_hien_tai - timedelta(days=1)
ngay_dau_thang = date(ngay_hien_tai.year, ngay_hien_tai.month, 1)
sql_pl1_1 = """
            WITH ALL_DAYS AS (   
                SELECT TO_DATE(:year || '-01-01', 'yyyy-mm-dd') + LEVEL - 1 AS DAY_DATE
                FROM DUAL
                CONNECT BY LEVEL <= TO_DATE(:year || '-12-31', 'yyyy-mm-dd') - TO_DATE(:year || '-01-01', 'yyyy-mm-dd') + 1
            )
            SELECT 
                TO_CHAR(DAY_DATE,'MM'),
                NVL(SUM(p.REVENUE), 0) AS REVENUE
            FROM 
                ALL_DAYS a
            LEFT JOIN 
                REVENUE_VAS p
            ON 
                TRUNC(p.CREATED_AT) = a.DAY_DATE
            GROUP BY 
                a.DAY_DATE
            ORDER BY 
                a.DAY_DATE
            """
sql_pl1_2 = """
            WITH FIRST_DAYS AS (
                SELECT ADD_MONTHS(TRUNC(TO_DATE(:year || '-01-01', 'yyyy-mm-dd'), 'YYYY'), LEVEL - 1) AS FIRST_DAY
                FROM DUAL
                CONNECT BY LEVEL <= 12
            )
            SELECT 
                NVL(SUM(p.REVENUE), 0) AS REVENUE,
                CASE 
                    WHEN NVL(SUM(p.ABONNEMENT), 0) = 0 THEN 0
                    ELSE NVL(SUM(p.REVENUE), 0) / NVL(SUM(p.ABONNEMENT), 0)
                END AS ARPU
            FROM 
                FIRST_DAYS a
            LEFT JOIN 
                REVENUE_VAS p
            ON 
                TRUNC(p.CREATED_AT,'MM') = a.FIRST_DAY
            GROUP BY 
                a.FIRST_DAY
            ORDER BY 
                a.FIRST_DAY 
            """
sql_pl2 = """
SELECT * 
FROM SERVICE_REVENUE_DAILY_DATA_EXPORT srdde
ORDER BY 
  SERVICE_NAME,
  CASE
    WHEN SPECIFIC_DAY IN ('Lũy kế Tháng', 'Lũy kế tới ngày') THEN 32
    WHEN REGEXP_LIKE(SPECIFIC_DAY, '^\d+$') THEN TO_NUMBER(SPECIFIC_DAY)
    ELSE NULL
  END
"""
sql_pl3_1 = """
    WITH SERVICES AS (
    SELECT DISTINCT(
    CASE
    	WHEN SERVICE_NAME IN ('FISH_BUY','FISH_HUNTER')
        THEN 'FISH_BUY_HUNTER'
    ELSE CAST(SERVICE_NAME AS VARCHAR2(100))
    END)  AS SERVICE_NAME
    FROM VAS_SERVICES WHERE STATUS = 1 
    ORDER BY SERVICE_NAME
    
),
DATES AS (
    SELECT 
        TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') + LEVEL - 1 AS DAY_DATE
    FROM DUAL
    CONNECT BY LEVEL <= LAST_DAY(TO_DATE(:input_date, 'YYYY-MM-DD')) - TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') + 1
)
SELECT 
    TO_CHAR(D.DAY_DATE, 'DD') day_,
    S.SERVICE_NAME,
    NVL(SUM(B.QUANTITY), 0) AS QUANTITY
FROM 
    DATES D
CROSS JOIN 
    SERVICES S
LEFT JOIN 
    (SELECT 
         CASE 
             WHEN SERVICE_NAME IN ('FISH_BUY', 'FISH_HUNTER') THEN 'FISH_BUY_HUNTER'
             ELSE CAST(SERVICE_NAME AS VARCHAR2(100))
         END AS SERVICE_NAME, 
         QUANTITY, 
         CREATED_AT 
     FROM ACCUMULATE_BY_DAY 
     WHERE TRUNC(CREATED_AT) BETWEEN TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') 
                                 AND LAST_DAY(TO_DATE(:input_date, 'YYYY-MM-DD'))) B
ON 
    S.SERVICE_NAME = B.SERVICE_NAME AND D.DAY_DATE = TRUNC(B.CREATED_AT)
GROUP BY 
    D.DAY_DATE, S.SERVICE_NAME
ORDER BY 
    D.DAY_DATE, S.SERVICE_NAME
            """
sql_pl3_2 = """
WITH SERVICES AS (
    SELECT DISTINCT(
    CASE
    	WHEN SERVICE_NAME IN ('FISH_BUY','FISH_HUNTER')
        THEN 'FISH_BUY_HUNTER'
    ELSE CAST(SERVICE_NAME AS VARCHAR2(100))
    END)  AS SERVICE_NAME
    FROM VAS_SERVICES WHERE STATUS = 1 
    ORDER BY SERVICE_NAME
    
),
DATES AS (
    SELECT 
        TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') + LEVEL - 1 AS DAY_DATE
    FROM DUAL
    CONNECT BY LEVEL <= LAST_DAY(TO_DATE(:input_date, 'YYYY-MM-DD')) - TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') + 1
)
SELECT 
    TO_CHAR(D.DAY_DATE, 'DD') day_,
    S.SERVICE_NAME,
    NVL(SUM(B.QUANTITY), 0) AS QUANTITY
FROM 
    DATES D
CROSS JOIN 
    SERVICES S
LEFT JOIN 
    (SELECT 
         CASE 
             WHEN SERVICE_NAME IN ('FISH_BUY', 'FISH_HUNTER') THEN 'FISH_BUY_HUNTER'
             ELSE CAST(SERVICE_NAME AS VARCHAR2(100))
         END AS SERVICE_NAME, 
         QUANTITY, 
         CREATED_AT 
     FROM NEW_REGISTER_BY_DAY 
     WHERE TRUNC(CREATED_AT) BETWEEN TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') 
                                 AND LAST_DAY(TO_DATE(:input_date, 'YYYY-MM-DD'))) B
ON 
    S.SERVICE_NAME = B.SERVICE_NAME AND D.DAY_DATE = TRUNC(B.CREATED_AT)
GROUP BY 
    D.DAY_DATE, S.SERVICE_NAME
ORDER BY 
    D.DAY_DATE, S.SERVICE_NAME
"""
sql_pl3_3 = """
WITH SERVICES AS (
    SELECT DISTINCT(
    CASE
    	WHEN SERVICE_NAME IN ('FISH_BUY','FISH_HUNTER')
        THEN 'FISH_BUY_HUNTER'
    ELSE CAST(SERVICE_NAME AS VARCHAR2(100))
    END)  AS SERVICE_NAME
    FROM VAS_SERVICES WHERE STATUS = 1 
    ORDER BY SERVICE_NAME
    
),
DATES AS (
    SELECT 
        TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') + LEVEL - 1 AS DAY_DATE
    FROM DUAL
    CONNECT BY LEVEL <= LAST_DAY(TO_DATE(:input_date, 'YYYY-MM-DD')) - TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') + 1
)
SELECT 
    TO_CHAR(D.DAY_DATE, 'DD') day_,
    S.SERVICE_NAME,
    NVL(SUM(B.QUANTITY), 0) AS QUANTITY
FROM 
    DATES D
CROSS JOIN 
    SERVICES S
LEFT JOIN 
    (SELECT 
         CASE 
             WHEN SERVICE_NAME IN ('FISH_BUY', 'FISH_HUNTER') THEN 'FISH_BUY_HUNTER'
             ELSE CAST(SERVICE_NAME AS VARCHAR2(100))
         END AS SERVICE_NAME, 
         QUANTITY, 
         CREATED_AT 
     FROM CANCLE_BY_DAY 
     WHERE TRUNC(CREATED_AT) BETWEEN TRUNC(TO_DATE(:input_date, 'YYYY-MM-DD'), 'MM') 
                                 AND LAST_DAY(TO_DATE(:input_date, 'YYYY-MM-DD'))) B
ON 
    S.SERVICE_NAME = B.SERVICE_NAME AND D.DAY_DATE = TRUNC(B.CREATED_AT)
GROUP BY 
    D.DAY_DATE, S.SERVICE_NAME
ORDER BY 
    D.DAY_DATE, S.SERVICE_NAME

"""
sql_pl4 = """
WITH CHANNELS_MPI AS (
    SELECT CHANNEL FROM VAS_CHANNELS WHERE STATUS = 1 
)
SELECT 
    c.CHANNEL,
    NVL(t1.QUANTITY, 0) AS QUANTITY_DATE,
    NVL(t2.QUANTITY, 0) AS QUANTITY_DATE_MINUS_1,
    NVL(t1.QUANTITY, 0) - NVL(t2.QUANTITY, 0) AS DELTA_QUANTITY,
    CASE 
        WHEN NVL(t2.QUANTITY, 0) = 0 THEN 
            CASE WHEN NVL(t1.QUANTITY, 0) = 0 THEN 0 ELSE 100 END
        ELSE 
            ROUND((NVL(t1.QUANTITY, 0) - NVL(t2.QUANTITY, 0)) * 100 / t2.QUANTITY, 2)
    END AS PERCENT_CHANGE,
    NVL(t3.QUANTITY, 0) AS QUANTITY_MONTH,
    NVL(t4.QUANTITY, 0) AS QUANTITY_MONTH_MINUS_1,
    NVL(t3.QUANTITY, 0) - NVL(t4.QUANTITY, 0) AS DELTA_QUANTITY_MONTH,
    CASE 
        WHEN NVL(t4.QUANTITY, 0) = 0 THEN 
            CASE WHEN NVL(t4.QUANTITY, 0) = 0 THEN 0 ELSE 100 END
        ELSE 
            ROUND((NVL(t3.QUANTITY, 0) - NVL(t4.QUANTITY, 0)) * 100 / t4.QUANTITY, 2)
    END AS PERCENT_CHANGE_MONTH
FROM 
    CHANNELS_MPI c
LEFT JOIN 
    (
        SELECT CHANNEL AS CHANNEL, sum(QUANTITY) AS QUANTITY
        FROM CHANNEL_REPORT_BY_DAY
        WHERE trunc(CREATED_AT) = to_date(:v_date,'yyyy-mm-dd') 
        GROUP BY channel
    ) t1
ON c.CHANNEL = t1.CHANNEL
LEFT JOIN 
    (
        SELECT CHANNEL AS CHANNEL,sum(QUANTITY) AS QUANTITY
        FROM CHANNEL_REPORT_BY_DAY
        WHERE trunc(CREATED_AT) = to_date(:v_date,'yyyy-mm-dd') - 1
        GROUP BY channel
    ) t2
ON c.CHANNEL = t2.CHANNEL
LEFT JOIN 
    (
        SELECT CHANNEL AS CHANNEL,sum(QUANTITY) AS QUANTITY
        FROM CHANNEL_REPORT_BY_DAY
        WHERE TRUNC(CREATED_AT, 'MM') = TRUNC(to_date(:v_date,'yyyy-mm-dd'), 'MM') 
        GROUP BY channel
    ) t3
ON c.CHANNEL = t3.CHANNEL
LEFT JOIN 
    (
        SELECT CHANNEL AS CHANNEL,sum(QUANTITY) AS QUANTITY
        FROM CHANNEL_REPORT_BY_DAY
        WHERE TRUNC(CREATED_AT, 'MM') = ADD_MONTHS(TRUNC(to_date(:v_date,'yyyy-mm-dd'), 'MM'), -1)
        GROUP BY channel
    ) t4
ON c.CHANNEL = t4.CHANNEL
"""


def call_package(all_packages):
    # Initialize OracleHook
    oracle_hook = OracleHook(oracle_conn_id='Mytel_DWH')
    try:
        for package_sql in all_packages:
            logger.info(f"Executing package: {package_sql}")
            oracle_hook.run(
                f"BEGIN {package_sql}(:ngay_tong_hop); END;",
                parameters={"ngay_tong_hop": ngay_tong_hop}
            )
    except Exception as e:
        logger.error(f"Error while executing package: {e}")
        raise
    
def call_package_pl2():
    # Initialize OracleHook
    oracle_hook = OracleHook(oracle_conn_id='Mytel_DWH')
    try:
        sql1 = "PCK_REPORT_VAS_REVENUE_DAILY.GET_REVENUE_DAILY_DATA_VAS"
        # sql2 ="PCK_REPORT_VAS_REVENUE_DAILY.GET_REVENUE_DAILY_DATA_EXPORT_TMP"
        sql3 = "PCK_REPORT_VAS_REVENUE_DAILY.GET_ALL_DAYS_SERVICES_EXPORT_TMP"
        sql4 = "PCK_REPORT_VAS_REVENUE_DAILY.GET_REVENUE_DAILY_DATA_EXPORT"
        logger.info("===============call pck pl2===================")
        oracle_hook.run(
            f"BEGIN {sql1}(:ngay_tong_hop); END;",
            parameters={"ngay_tong_hop": ngay_tong_hop.strftime("%Y-%m-%d")}
        )
        # oracle_hook.run(
        #     f"BEGIN {sql2}(:ngay_tong_hop); END;",
        #     parameters={"ngay_tong_hop": ngay_tong_hop.strftime("%Y-%m-%d")}
        # )
        oracle_hook.run(
            f"BEGIN {sql3}; END;",
        )
        oracle_hook.run(
            f"BEGIN {sql4}(:ngay_tong_hop); END;",
            parameters={"ngay_tong_hop": ngay_tong_hop.strftime("%Y-%m-%d")}
        )
        logger.info("===========call pck pl2 hoan thanh================")
    except Exception as e:
        logger.error(f"Error while executing package: {e}")
        raise

def file_excel_processing(**kwargs):
    try:
        conn = kwargs.get("connection")
        template_path = kwargs.get("template_path")
        pl = kwargs.get("pl")
        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        wb = openpyxl.load_workbook(template_path)
        if pl == 'PL01':
            logger.info("=========Xu ly phu luc 1=========")
            ws = wb.worksheets[0]
            ws.cell(row=2, column=1, value="Thời điểm xuất báo cáo: "
                    + f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            ws.cell(row=3, column=1, value="Từ ngày:  " + f"{ngay_dau_thang.day} tháng {ngay_dau_thang.month} năm {ngay_dau_thang.year}"
                    + f" Đến ngày: {ngay_tong_hop.day} tháng {ngay_tong_hop.month} năm {ngay_tong_hop.year}")
            ws.cell(row=5, column=4, value="Lũy kế tháng " + f"{ngay_tong_hop.month}")
            ws.cell(row=5, column=6, value="So với tháng " + f"{ngay_tong_hop.month - 1}")
            ws.cell(row=6, column=6, value="Lũy kế tháng " + f"{ngay_tong_hop.month - 1}")
            ws.cell(row=10, column=2, value="Dữ liệu năm " + f"{ngay_tong_hop.year}")
            ws.cell(row=26, column=2, value="Dữ liệu năm " + f"{ngay_tong_hop.year - 1}")
            with conn.cursor() as cursor:
                yearStr = str(ngay_tong_hop.year)
                cursor.execute(sql_pl1_1, {"year": yearStr,})
                results = cursor.fetchall()
                row_check = 13
                col_num = 3
                row_num = 13
                for row in results:
                    for i in range(2):
                        if i == 0:
                            month_map = {
                                '01': 13,
                                '02': 14,
                                '03': 15,
                                '04': 16,
                                '05': 17,
                                '06': 18,
                                '07': 19,
                                '08': 20,
                                '09': 21,
                                '10': 22,
                                '11': 23,
                                '12': 24
                            }
                            month = row[0]
                            row_num = month_map.get(month, 13)
                            if row_num > row_check:
                                row_check = row_num
                                col_num = 3
                        else:
                            cell = ws.cell(
                                row=row_num, column=col_num, value=row[i])
                            cell.border = thin_border
                            col_num += 1
            with conn.cursor() as cursor:
                yearStr = str(ngay_tong_hop.year)
                cursor.execute(sql_pl1_2, {"year": yearStr,})
                results = cursor.fetchall()
                row_num = 13
                for row in results:
                    col_num = 34
                    for value in row:
                        cell = ws.cell(
                            row=row_num, column=col_num, value=value)
                        cell.border = thin_border
                        col_num += 1
                    row_num += 1
            with conn.cursor() as cursor:
                yearStr = str(ngay_tong_hop.year - 1)
                cursor.execute(sql_pl1_1, {"year": yearStr,})
                results = cursor.fetchall()
                row_check = 29
                row_num = 29
                col_num = 3
                for row in results:
                    for i in range(2):
                        if i == 0:
                            month_map = {
                                '01': 29,
                                '02': 30,
                                '03': 31,
                                '04': 32,
                                '05': 33,
                                '06': 34,
                                '07': 35,
                                '08': 36,
                                '09': 37,
                                '10': 38,
                                '11': 39,
                                '12': 40
                            }
                            month = row[0]
                            row_num = month_map.get(month, 29)
                            if row_num > row_check:
                                row_check = row_num
                                col_num = 3
                        else:
                            cell = ws.cell(
                                row=row_num, column=col_num, value=row[i])
                            cell.border = thin_border
                            col_num += 1
            with conn.cursor() as cursor:
                yearStr = str(ngay_tong_hop.year - 1)
                cursor.execute(sql_pl1_2, {"year": yearStr,})
                results = cursor.fetchall()
                row_num = 29
                for row in results:
                    col_num = 34
                    for value in row:
                        cell = ws.cell(
                            row=row_num, column=col_num, value=value)
                        cell.border = thin_border
                        col_num += 1
                    row_num += 1
            logger.info("=========Xu ly phu luc 1 hoan thanh=========")
        if pl == 'PL02':
            logger.info("=========Xu ly phu luc 2=========")
            ws = wb.worksheets[0]
            ws.cell(row=2, column=1, value="Thời điểm xuất báo cáo: " 
                    + f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            ws.cell(row=3, column=1, value="Từ ngày:  " + f"{ngay_dau_thang.day} tháng {ngay_dau_thang.month} năm {ngay_dau_thang.year}"
                    + f" Đến ngày: {ngay_tong_hop.day} tháng {ngay_tong_hop.month} năm {ngay_tong_hop.year}")
            ws.cell(row=5, column=3, value="So sánh với tháng " + f"{ngay_tong_hop.month - 1} năm {ngay_tong_hop.year}")
            ws.cell(row=5, column=5, value="So sánh với tháng " + f"{ngay_tong_hop.month} năm {ngay_tong_hop.year - 1}")
            ws.cell(row=5, column=7, value="Năm " + f"{ngay_tong_hop.year}")
            ws.cell(row=5, column=19, value="Năm " + f"{ngay_tong_hop.year - 1}")
            with conn.cursor() as cursor:
                logger.info(
                    "=========Lay du lieu TONG HOP DOANH THU THEO DICH VU=========")
                with conn.cursor() as cursor:
                    cursor.execute(sql_pl2)
                    results = cursor.fetchall()
                    row_num = 7
                    for row in results:
                        col_num = 1
                        for value in row:
                            cell = ws.cell(
                                row=row_num, column=col_num, value=value)
                            cell.border = thin_border
                            col_num += 1
                        row_num += 1
            logger.info("=========Xu ly phu luc 2 hoan thanh=========")
        if pl == 'PL03':
            logger.info("=========Xu ly phu luc 3=========")
            wb = openpyxl.load_workbook(template_path)
            ws0 = wb.worksheets[0]
            ws1 = wb.worksheets[1]
            ws2 = wb.worksheets[2]
            tieuDe = 'BÁO CÁO SỐ LIỆU THUÊ BAO LŨY KẾ ĐẾN NGÀY ' + f"{ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}"
            ws0.cell(row=1, column=3, value=tieuDe)
            ws0.cell(row=2, column=1, value="Thời điểm xuất báo cáo: " 
                     + f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            cellThang = ws0.cell(row=5, column=2, value=f"Tháng {ngay_hien_tai.month}")
            cellThang.border = thin_border
            with conn.cursor() as cursor:
                logger.info("=========Lay du lieu luy ke thang=========")
                cursor.execute(sql_pl3_1, {"input_date": ngay_tong_hop.strftime("%Y-%m-%d")})
                result = cursor.fetchall()
                row_num = 7
                col_num = 2
                ngay = 1
                for value in result:                    
                    if value[0] == '01':
                        cellS = ws0.cell(row=row_num,column=1,value=value[1])
                        cellS.border = thin_border
                        cell = ws0.cell(row=row_num,column=2,value=value[2])
                        cell.border = thin_border
                        row_num += 1
                    else:
                        ngay_m = int(value[0])
                        if ngay_m > ngay:
                            ngay = ngay_m
                            col_num += 1
                            row_num = 7
                        cell = ws0.cell(row=row_num,column=col_num,value=value[2])
                        cell.border = thin_border
                        row_num += 1    
                logger.info(
                    "=========Lay du lieu luy ke thang hoan thanh=========")
            #################
            tieuDe = 'BÁO CÁO SỐ LIỆU THUÊ BAO ĐĂNG KÝ MỚI ĐẾN NGÀY ' + \
                f"{ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}"
            ws1.cell(row=1, column=3, value=tieuDe)
            ws1.cell(row=2, column=1, value="Thời điểm xuất báo cáo: " 
                     + f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            cellThang = ws1.cell(row=5, column=2, value=f"Tháng {ngay_hien_tai.month}")
            cellThang.border = thin_border
            with conn.cursor() as cursor:
                logger.info("=========Lay du lieu luy ke thang=========")
                cursor.execute(sql_pl3_2, {"input_date": ngay_tong_hop.strftime("%Y-%m-%d")})
                result = cursor.fetchall()
                row_num = 7
                col_num = 2
                ngay = 1
                for value in result:                    
                    if value[0] == '01':
                        cellS = ws1.cell(row=row_num,column=1,value=value[1])
                        cellS.border = thin_border
                        cell = ws1.cell(row=row_num,column=2,value=value[2])
                        cell.border = thin_border
                        row_num += 1
                    else:
                        ngay_m = int(value[0])
                        if ngay_m > ngay:
                            ngay = ngay_m
                            col_num += 1
                            row_num = 7
                        cell = ws1.cell(row=row_num,column=col_num,value=value[2])
                        cell.border = thin_border
                        row_num += 1    
                logger.info(
                    "=========Lay du lieu dang ky moi thang hoan thanh=========")
            #################
            tieuDe = 'BÁO CÁO SỐ LIỆU THUÊ BAO HỦY ĐẾN NGÀY ' + \
                f"{ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}"
            ws2.cell(row=1, column=3, value=tieuDe)
            ws2.cell(row=2, column=1, value="Thời điểm xuất báo cáo: " 
                     + f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            cellThang = ws2.cell(row=5, column=2, value=f"Tháng {ngay_hien_tai.month}")
            cellThang.border = thin_border
            with conn.cursor() as cursor:
                logger.info("=========Lay du lieu luy ke thang=========")
                cursor.execute(sql_pl3_1, {"input_date": ngay_tong_hop.strftime("%Y-%m-%d")})
                result = cursor.fetchall()
                row_num = 7
                col_num = 2
                ngay = 1
                for value in result:                    
                    if value[0] == '01':
                        cellS = ws2.cell(row=row_num,column=1,value=value[1])
                        cellS.border = thin_border
                        cell = ws2.cell(row=row_num,column=2,value=value[2])
                        cell.border = thin_border
                        row_num += 1
                    else:
                        ngay_m = int(value[0])
                        if ngay_m > ngay:
                            ngay = ngay_m
                            col_num += 1
                            row_num = 7
                        cell = ws2.cell(row=row_num,column=col_num,value=value[2])
                        cell.border = thin_border
                        row_num += 1    
                logger.info(
                    "=========Lay du lieu HUY thang hoan thanh=========")
            logger.info("=========Xu ly phu luc 3 hoan thanh=========")
        if pl == 'PL04':
            logger.info("=========Xu ly phu luc 4=========")
            wb = openpyxl.load_workbook(template_path)
            ws = wb.worksheets[0]
            tieuDe = 'BÁO CÁO THUÊ BAO KÊNH BÁN ' + \
                f"{ngay_hien_tai.month} năm {ngay_hien_tai.year}"
            ws.cell(row=1, column=3, value=tieuDe)
            ws.cell(row=2, column=1, value="Thời điểm xuất báo cáo: " 
                     + f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            ws.cell(row=7, column= 3, value=f"{ngay_tong_hop.day}-{ngay_tong_hop.month}-{ngay_tong_hop.year}")
            ws.cell(row=7, column= 4, value=f"{ngay_tong_hop.day - 1}-{ngay_tong_hop.month}-{ngay_tong_hop.year}")
            ws.cell(row=7, column= 7, value=f"Tháng {ngay_tong_hop.month} năm {ngay_tong_hop.year}")
            ws.cell(row=7, column= 8, value=f"Tháng {ngay_tong_hop.month - 1} năm {ngay_tong_hop.year}")
            with conn.cursor() as cursor:
                cursor.execute(sql_pl4, {"v_date": ngay_tong_hop.strftime("%Y-%m-%d")})
                results = cursor.fetchall()
                row_num = 8
                tt = 1
                for row in results:
                    col_num = 2
                    for value in row:
                        cell = ws.cell(
                            row=row_num, column=col_num, value=value)
                        cell.border = thin_border
                        col_num += 1
                    cellTT = ws.cell(row=row_num, column=1, value=tt)
                    cellTT.border = thin_border
                    tt += 1
                    row_num += 1
            logger.info("=========Xu ly phu luc 4=========")

        ######################
        # Lấy tên file từ đường dẫn
        base_name = os.path.basename(pl + "_" + ngay_tong_hop.strftime("%d%m%Y") + ".xlsx")
        new_directory = "./export/"  # Đường dẫn thư mục bạn muốn lưu
        new_file_path = os.path.join(new_directory, base_name)
        wb.save(new_file_path)
    except Exception as e:
        logger.error(e)


def export_data():
    try:
        oracle_hook = OracleHook(oracle_conn_id='Mytel_DWH')
        conn = oracle_hook.get_conn()
        logger.info(
            "===========Bat dau tien trinh xu ly file excel===============")
        
        lstTemplatePath = ['./temp/PL01.xlsx', './temp/PL02.xlsx',
                           './temp/PL03.xlsx', './temp/PL04.xlsx']
        
        for path in lstTemplatePath:
            if (path == './temp/PL01.xlsx'):
                pl = "PL01"
            if (path == './temp/PL02.xlsx'):
                pl = "PL02"
            if (path == './temp/PL03.xlsx'):
                pl = "PL03"
            if (path == './temp/PL04.xlsx'):
                pl = "PL04"
            op_kwargs = {
                "connection": conn,
                "template_path": path,
                "pl": pl
            }
            file_excel_processing(**op_kwargs)
        logger.info(
            "===========Hoan thanh tien trinh xu ly file excel===============")
    except Exception as e:
        logger.exception(e)
    finally:
        try:
            if conn:
                conn.close()
        except NameError:
            pass


vas_report_daily_args = {
    "owner": "airflow",
    "depends_on_past": False,
    "start_date": datetime(2024, 12, 1),
    "retries": 0,
}

with DAG(
    dag_id="vas_report_daily",
    default_args=vas_report_daily_args,
    description="Tien trinh tong hop du lieu VAS thi truong Mytel",
    schedule_interval="0 9 * * *",  # Chạy lúc 9 giờ sáng mỗi ngày
    catchup=False,
) as dag:
    call_package_vas_daily = PythonOperator(
        task_id="call_package_vas_daily",
        python_callable=call_package,
        op_kwargs={
            "all_packages": ["PCK_REPORT_VAS.ACCUMULATE", 
                             "PCK_REPORT_VAS.NEW_REGISTER",
                             "PCK_REPORT_VAS.CANCEL",
                             "PCK_REPORT_VAS.REPORT_BY_CHANNEL",
                             "PCK_REPORT_VAS.REVENUE_VAS"
                             ]
        }
    )
    
    call_package_vas_daily_pl2 = PythonOperator(
        task_id="call_package_vas_daily_pl2",
        python_callable=call_package_pl2,
    )

    export_data_vas_daily = PythonOperator(
        task_id="export_data_vas_daily",
        python_callable=export_data,
    )
    
    call_package_vas_daily >> call_package_vas_daily_pl2 >> export_data_vas_daily
