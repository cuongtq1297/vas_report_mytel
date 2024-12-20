from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.operators.email import EmailOperator
import cx_Oracle
import openpyxl
from openpyxl.styles import Border,Side
import logging
import os
import calendar
from datetime import datetime, date

# Khởi tạo logger
logger = logging.getLogger("airflow.task")
logger.setLevel(logging.INFO)
sql_pl1_1 = """
            WITH ALL_DAYS AS (   
                SELECT TO_DATE('2024' || '-01-01', 'yyyy-mm-dd') + LEVEL - 1 AS DAY_DATE
                FROM DUAL
                CONNECT BY LEVEL <= TO_DATE('2024' || '-12-31', 'yyyy-mm-dd') - TO_DATE('2024' || '-01-01', 'yyyy-mm-dd') + 1
            )
            SELECT 
                TO_CHAR(DAY_DATE,'MM'),
                NVL(SUM(p.REVENUE), 0) AS REVENUE
            FROM 
                ALL_DAYS a
            LEFT JOIN 
                REVENUE_VAS p
            ON 
                TRUNC(p.CREATED_AT) = a.DAY_DATE
            GROUP BY 
                a.DAY_DATE
            ORDER BY 
                a.DAY_DATE
            """
sql_pl1_2 = """
                WITH ALL_DAYS AS (   
                SELECT TO_DATE('2023' || '-01-01', 'yyyy-mm-dd') + LEVEL - 1 AS DAY_DATE
                FROM DUAL
                CONNECT BY LEVEL <= TO_DATE('2023' || '-12-31', 'yyyy-mm-dd') - TO_DATE('2023' || '-01-01', 'yyyy-mm-dd') + 1
            )
            SELECT 
                TO_CHAR(DAY_DATE,'MM'),
                NVL(SUM(p.REVENUE), 0) AS REVENUE
            FROM 
                ALL_DAYS a
            LEFT JOIN 
                REVENUE_VAS p
            ON 
                TRUNC(p.CREATED_AT) = a.DAY_DATE
            GROUP BY 
                a.DAY_DATE
            ORDER BY 
                a.DAY_DATE
            """
sql_pl1_3 = """
            WITH FIRST_DAYS AS (
                SELECT ADD_MONTHS(TRUNC(TO_DATE('2024' || '-01-01', 'yyyy-mm-dd'), 'YYYY'), LEVEL - 1) AS FIRST_DAY
                FROM DUAL
                CONNECT BY LEVEL <= 12
            )
            SELECT 
                NVL(SUM(p.REVENUE), 0) AS REVENUE,
                NVL(SUM(p.REVENUE), 0) / NVL(SUM(p.ABONNEMENT), 0) AS ARPU
            FROM 
                FIRST_DAYS a
            LEFT JOIN 
                REVENUE_VAS p
            ON 
                TRUNC(p.CREATED_AT,'MM') = a.FIRST_DAY
            GROUP BY 
                a.FIRST_DAY
            ORDER BY 
                a.FIRST_DAY 
            """
sql_pl1_4 = """
            WITH FIRST_DAYS AS (
                SELECT ADD_MONTHS(TRUNC(TO_DATE('2024' || '-01-01', 'yyyy-mm-dd'), 'YYYY'), LEVEL - 1) AS FIRST_DAY
                FROM DUAL
                CONNECT BY LEVEL <= 12
            )
            SELECT 
                NVL(SUM(p.REVENUE), 0) AS REVENUE,
                NVL(SUM(p.REVENUE), 0) / NVL(SUM(p.ABONNEMENT), 0) AS ARPU
            FROM 
                FIRST_DAYS a
            LEFT JOIN 
                REVENUE_VAS p
            ON 
                TRUNC(p.CREATED_AT,'MM') = a.FIRST_DAY
            GROUP BY 
                a.FIRST_DAY
            ORDER BY 
                a.FIRST_DAY 
            """                         
sql_pl2 = "SELECT * FROM service_revenua_daily_data_export"                    
sql_pl3_1 = "SELECT * FROM ACCUM_DATA_EXPORT"
sql_pl3_2 = "SELECT * FROM NEW_REGISTER_DATA_EXPORT"
sql_pl3_3 = "SELECT * FROM CANCEL_DATA_EXPORT"
sql_pl4 = """
            SELECT CHANNEL, QUANTITY_DATE, QUANTITY_DATE_MINUS_1, DELTA_QUANTITY, PERCENT_CHANGE, QUANTITY_MONTH, QUANTITY_MONTH_MINUS_1, DELTA_QUANTITY_MONTH, PERCENT_CHANGE_MONTH
            FROM MYTEL_WAREHOUSE_UAT.REPORT_BY_CHANNEL_DATA_EXPORT
            """
            
# Hàm gọi package Oracle
def call_oracle_package():
    dsn = cx_Oracle.makedsn("10.201.214.40", "1521", service_name="mytel") 
    conn = cx_Oracle.connect(user="MYTEL_WAREHOUSE_UAT", password="MYTEL_WAREHOUSE_UAT#2023", dsn=dsn)
    try:
        with conn.cursor() as cursor:
            # Gọi package/procedure
            v_date = date(2024, 11, 6) 
            cursor.callproc("PCK_TOTAL_VAS.TOTAL_MYGO_VAS",[input_date])
            # cursor.callproc("PCK_REPORT_VAS.GET_ACCUM_DATA_EXPORT",[v_date])
            conn.commit()  # Commit nếu có tác động dữ liệu
    except cx_Oracle.DatabaseError as e:
        logging(f"Error: {e}")
        raise
    finally:
        conn.close()
# Ham xu ly file excel
def get_data_and_export_file(**kwargs):
    try:
        conn = kwargs.get("connection")
        template_path = kwargs.get("template_path")
        pl = kwargs.get("pl")
        # Lấy ngày hiện tại
        ngay_hien_tai = datetime.now()
        so_ngay_trong_thang = calendar.monthrange(ngay_hien_tai.year, ngay_hien_tai.month)[1]
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        wb = openpyxl.load_workbook(template_path)
        if pl == 'PL01':
            logger.info("=========Xu ly phu luc 1=========")
            ws = wb.worksheets[0]
            ws.cell(row=2, column=1, value="Thời điểm xuất báo cáo: "+ f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            ws.cell(row=3, column=1, value="Từ ngày  "+ f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            with conn.cursor() as cursor:
                cursor.execute(sql_pl1_1)
                results = cursor.fetchall()
                row_check = 13
                col_num = 3 
                row_num = 13
                for row in results: 
                    for i in range(2):
                        if i == 0:
                            month_map = {
                                '01': 13,
                                '02': 14,
                                '03': 15,
                                '04': 16,
                                '05': 17,
                                '06': 18,
                                '07': 19,
                                '08': 20,
                                '09': 21,
                                '10': 22,
                                '11': 23,
                                '12': 24
                            }
                            month = row[0]
                            row_num = month_map.get(month, 13)   
                            if row_num > row_check:
                                row_check = row_num
                                col_num = 3
                        else:
                            cell = ws.cell(row=row_num, column=col_num, value=row[i])
                            cell.border = thin_border
                            col_num += 1 
            with conn.cursor() as cursor:
                cursor.execute(sql_pl1_3)
                results = cursor.fetchall()
                row_num = 13
                for row in results: 
                    col_num = 34 
                    for value in row:
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = thin_border
                        col_num += 1
                    row_num += 1        
            with conn.cursor() as cursor:
                cursor.execute(sql_pl1_2)
                results = cursor.fetchall()
                row_check = 29
                row_num = 29
                col_num = 3
                for row in results:
                    for i in range(2):
                        if i == 0:
                            month_map = {
                                '01': 29,
                                '02': 30,
                                '03': 31,
                                '04': 32,
                                '05': 33,
                                '06': 34,
                                '07': 35,
                                '08': 36,
                                '09': 37,
                                '10': 38,
                                '11': 39,
                                '12': 40
                            }
                            month = row[0]
                            row_num = month_map.get(month, 29)   
                            if row_num > row_check:
                                row_check = row_num
                                col_num = 3
                        else:
                            cell = ws.cell(row=row_num, column=col_num, value=row[i])
                            cell.border = thin_border
                            col_num += 1
            with conn.cursor() as cursor:
                cursor.execute(sql_pl1_4)
                results = cursor.fetchall()
                row_num = 29
                for row in results: 
                    col_num = 34 
                    for value in row:
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = thin_border
                        col_num += 1
                    row_num += 1                   
            logger.info("=========Xu ly phu luc 1 hoan thanh=========")                
        if pl == 'PL02':
            logger.info("=========Xu ly phu luc 2=========")
            ws = wb.worksheets[0]
            ws.cell(row=2, column=1, value="Thời điểm xuất báo cáo: "+ f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            ws.cell(row=5, column=3, value="So sánh với tháng 06-2024") #fix de test
            ws.cell(row=5, column=5, value="So sánh với tháng 06-2023") #fix de test
            with conn.cursor() as cursor:
                logger.info("=========Lay du lieu TONG HOP DOANH THU THEO DICH VU=========")
                with conn.cursor() as cursor:
                    cursor.execute(sql_pl2)
                    results = cursor.fetchall()
                    row_num = 7
                    for row in results:
                        col_num = 1 
                        for value in row:
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            cell.border = thin_border
                            col_num += 1
                        row_num += 1
            logger.info("=========Xu ly phu luc 2 hoan thanh=========")
        if pl == 'PL03':
            logger.info("=========Xu ly phu luc 3=========")
            wb = openpyxl.load_workbook(template_path)
            ws0 = wb.worksheets[0]
            ws1 = wb.worksheets[1]
            ws2 = wb.worksheets[2]
            delta = 31 - so_ngay_trong_thang
            # delta = 1 # de test
            # if delta > 0:
            #     # xóa bứt cột nếu tháng ít hơn 31 ngày
            #     for i in range(delta):
            #         # ws0.delete_cols(so_ngay_trong_thang + 2)
            #         ws0.delete_cols(30 + 2) # de test
            # Tiêu đề và các ô cần thêm ngày.
            tieuDe = 'BÁO CÁO SỐ LIỆU THUÊ BAO LŨY KẾ ĐẾN NGÀY ' + f"{ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}"
            ws0.cell(row=1, column=3, value=tieuDe)
            ws0.cell(row=2, column=1, value="Thời điểm xuất báo cáo: "+ f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            cellThang = ws0.cell(row=5, column=2, value=f"Tháng {ngay_hien_tai.month}")
            cellThang.border = thin_border
            with conn.cursor() as cursor:
                logger.info("=========Lay du lieu luy ke thang=========")
                cursor.execute(sql_pl3_1)
                result = cursor.fetchall()
                row_num = 7 
                col_num = 2  
                num_columns = len(cursor.description); 
                for value in result:
                    for i in range(num_columns): 
                        cell = ws0.cell(row=row_num, column=col_num, value=value[i])
                        cell.border = thin_border
                        row_num += 1
                    row_num = 7    
                    col_num += 1
                logger.info("=========Lay du lieu luy ke thang hoan thanh=========")    
            #################
            tieuDe = 'BÁO CÁO SỐ LIỆU THUÊ BAO ĐĂNG KÝ MỚI ĐẾN NGÀY ' + f"{ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}"
            ws1.cell(row=1, column=3, value=tieuDe)
            ws1.cell(row=2, column=1, value="Thời điểm xuất báo cáo: "+ f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            cellThang = ws1.cell(row=5, column=2, value=f"Tháng {ngay_hien_tai.month}")
            cellThang.border = thin_border
            with conn.cursor() as cursor:
                logger.info("=========Lay du lieu dang ky moi thang=========")
                cursor.execute(sql_pl3_2)
                result = cursor.fetchall()
                row_num = 7 
                col_num = 2  
                num_columns = len(cursor.description); 
                for value in result:
                    for i in range(num_columns): 
                        cell = ws1.cell(row=row_num, column=col_num, value=value[i])
                        cell.border = thin_border
                        row_num += 1
                    row_num = 7    
                    col_num += 1
                logger.info("=========Lay du lieu dang ky moi thang hoan thanh=========")    
            #################
            tieuDe = 'BÁO CÁO SỐ LIỆU THUÊ BAO HỦY ĐẾN NGÀY ' + f"{ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}"
            ws2.cell(row=1, column=3, value=tieuDe)
            ws2.cell(row=2, column=1, value="Thời điểm xuất báo cáo: "+ f"Ngày {ngay_hien_tai.day} tháng {ngay_hien_tai.month} năm {ngay_hien_tai.year}")
            cellThang = ws2.cell(row=5, column=2, value=f"Tháng {ngay_hien_tai.month}")
            cellThang.border = thin_border
            with conn.cursor() as cursor:
                logger.info("=========Lay du lieu HUY thang=========")
                cursor.execute(sql_pl3_3)
                result = cursor.fetchall()
                row_num = 7 
                col_num = 2  
                num_columns = len(cursor.description); 
                for value in result:
                    for i in range(num_columns): 
                        cell = ws2.cell(row=row_num, column=col_num, value=value[i])
                        cell.border = thin_border
                        row_num += 1
                    row_num = 7    
                    col_num += 1
                logger.info("=========Lay du lieu HUY thang hoan thanh=========")
            logger.info("=========Xu ly phu luc 3 hoan thanh=========")        
        if pl == 'PL04':
            logger.info("=========Xu ly phu luc 4=========")
            wb = openpyxl.load_workbook(template_path)
            ws = wb.worksheets[0]
            with conn.cursor() as cursor:
                cursor.execute(sql_pl4)
                results = cursor.fetchall()
                row_num = 8
                tt = 1
                for row in results:
                    col_num = 2
                    for value in row:
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = thin_border
                        col_num += 1
                    cellTT = ws.cell(row=row_num, column=1, value=tt) 
                    cellTT.border = thin_border
                    tt += 1     
                    row_num += 1 
            logger.info("=========Xu ly phu luc 4=========")
            
        ######################           
        base_name = os.path.basename(template_path)  # Lấy tên file từ đường dẫn
        new_directory = "./export/"  # Đường dẫn thư mục bạn muốn lưu
        new_file_path = os.path.join(new_directory, base_name)
        wb.save(new_file_path)   
    except Exception as e:
        logger.error(e);        
    

                
def exportr_to_excel(**kwargs):
    try:     
        logger.info("===========Bat dau tien trinh xu ly file excel===============")
        dsn = cx_Oracle.makedsn("10.201.214.40", "1521", service_name="mytel")  # Cấu hình DSN
        conn = cx_Oracle.connect(user="MYTEL_WAREHOUSE_UAT", password="MYTEL_WAREHOUSE_UAT#2023", dsn=dsn)
        lstTemplatePath = ['./temp/PL01.xlsx','./temp/PL02.xlsx','./temp/PL03.xlsx','./temp/PL04.xlsx']
        for path in lstTemplatePath:
            if (path == './temp/PL01.xlsx'):
                pl = "PL01"
            if (path == './temp/PL02.xlsx'):
                pl = "PL02"
            if (path == './temp/PL03.xlsx'):
                pl = "PL03"
            if (path == './temp/PL04.xlsx'):
                pl = "PL04"    
            op_kwargs = {
                            "connection": conn,
                            "template_path": path,
                            "pl": pl
                        }
            get_data_and_export_file(**op_kwargs)
        logger.info("===========Hoan thanh tien trinh xu ly file excel===============")    
    except cx_Oracle.DatabaseError as e:
        logger.exception(e)
    finally:
        try: 
            if conn: conn.close() 
        except NameError: 
            pass    
default_args = {
    "owner": "airflow",
    "depends_on_past": False,
    "start_date": datetime(2024, 12, 1),
    "retries": 0,
}

with DAG(
    dag_id="call_oracle_package_with_date",
    default_args=default_args,
    description="Call Oracle Package with DATE parameter using PythonOperator",
    schedule_interval=None,  # Chạy thủ công
    catchup=False,
) as dag:

    # Task sử dụng PythonOperator
    call_package_task = PythonOperator(
        task_id="call_package_with_date",
        python_callable=call_oracle_package,
    )
    
    
    # Task export file file excel
    export_excel_report = PythonOperator(
        task_id="export_excel_report",
        python_callable=exportr_to_excel,
        op_kwargs={
            "input_date": date(2024, 10, 2),  # Truyền vào DATE (datetime.date object)
        },
    )
    
    # Gửi email
#     send_email = EmailOperator(
#     task_id='send_email',
#     to='cuongtq1297@gmail.com',  # Người nhận email
#     subject='Test Email with Attachment from Airflow',
#     html_content='<h3>This is a test email with an attachment sent from Airflow</h3>',
#     files=['./export/PL01.xlsx'],  # Đường dẫn tới file đính kèm
#     dag=dag1
0
# )

call_package_task >> export_excel_report
       